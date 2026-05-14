"""
Excel fayldan darslarni parse qilish.
Har bir sheet = bir guruh jadvali.
Struktura: col0=sana, col1=hafta_kuni, col2=para, col3=vaqt, col4=modul, col5=tur, col6=oqituvchi, col7=xona
"""
import pandas as pd
from datetime import date

SKIP_SHEETS = ['хафталик', 'Лист1', 'Лист2', 'Лист4', 'МАРТ TAQSIMOT']

DAY_MAP = {
    'dushanba': 'Dushanba', 'душанба': 'Dushanba',
    'seshanba': 'Seshanba', 'сешанба': 'Seshanba',
    'chorshanba': 'Chorshanba', 'чоршанба': 'Chorshanba',
    'payshanba': 'Payshanba', 'пайшанба': 'Payshanba',
    'juma': 'Juma', 'жума': 'Juma',
    'shanba': 'Shanba', 'шанба': 'Shanba',
}

def parse_excel(file_path):
    """
    Returns: list of dicts with keys: sheet_name, lessons
    lessons: list of dicts: sana, hafta_kuni, para, vaqt, modul, tur, oqituvchi, xona
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        sheets = [s for s in wb.sheetnames if s not in SKIP_SHEETS]
    except Exception as e:
        raise ValueError(f"Excel fayl o'qishda xato: {e}")

    results = []
    for sheet_name in sheets:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            lessons = _parse_sheet(df)
            results.append({
                'sheet_name': sheet_name,
                'lessons': lessons,
                'count': len(lessons)
            })
        except Exception as e:
            results.append({'sheet_name': sheet_name, 'lessons': [], 'error': str(e)})
    return results


def _parse_sheet(df):
    lessons = []
    current_date = None
    current_day = None

    for idx, row in df.iterrows():
        if idx < 3:
            continue

        col0 = row.iloc[0] if len(row) > 0 and pd.notna(row.iloc[0]) else None
        col1 = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
        col2 = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ''
        col3 = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else ''
        col4 = str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else ''
        col5 = str(row.iloc[5]).strip() if len(row) > 5 and pd.notna(row.iloc[5]) else ''
        col6 = str(row.iloc[6]).strip() if len(row) > 6 and pd.notna(row.iloc[6]) else ''
        col7 = str(row.iloc[7]).strip() if len(row) > 7 and pd.notna(row.iloc[7]) else ''

        # Date cell
        if col0 is not None and hasattr(col0, 'strftime'):
            current_date = col0.date() if hasattr(col0, 'date') else col0
            current_day = DAY_MAP.get(col1.lower(), col1)
        elif col0 is not None and isinstance(col0, str) and col0.strip():
            pass  # string in col0 = not a date row

        # Lesson row: has time and module
        if (current_date and col3 and col3 != 'nan' and
                col4 and col4 != 'nan' and
                '___' not in col4 and
                'Modul' not in col4 and
                'vaqt' not in col4.lower()):
            lessons.append({
                'sana': current_date,
                'hafta_kuni': current_day or '',
                'para': col2 if col2 != 'nan' else '',
                'vaqt': col3,
                'modul': col4,
                'tur': col5 if col5 != 'nan' else '',
                'oqituvchi': col6 if col6 != 'nan' else '',
                'xona': col7 if col7 != 'nan' else '',
            })
    return lessons
