from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import JsonResponse
from datetime import date, timedelta, datetime
import os, tempfile

from .models import Guruh, Dars
from .excel_parser import parse_excel

HAFTA_KUNLARI = ['Dushanba', 'Seshanba', 'Chorshanba', 'Payshanba', 'Juma', 'Shanba']

# ─── MOBILE ───────────────────────────────────────────────

def mobile_home(request):
    guruhlar = Guruh.objects.all()
    selected_guruh = None
    week_schedule = []
    weeks = []
    selected_week = None
    today = date.today()

    guruh_id = request.GET.get('guruh')
    if guruh_id:
        selected_guruh = get_object_or_404(Guruh, id=guruh_id)
        weeks = selected_guruh.get_all_weeks()

        week_param = request.GET.get('hafta')
        if week_param:
            try:
                selected_week = datetime.strptime(week_param, '%Y-%m-%d').date()
            except Exception:
                selected_week = None
        if not selected_week:
            cur_mon = today - timedelta(days=today.weekday())
            selected_week = cur_mon if cur_mon in weeks else (weeks[0] if weeks else None)

        if selected_week:
            for i in range(6):
                day_date = selected_week + timedelta(days=i)
                lessons = list(Dars.objects.filter(guruh=selected_guruh, sana=day_date).order_by('para', 'vaqt'))
                week_schedule.append({
                    'date': day_date,
                    'day_name': HAFTA_KUNLARI[i],
                    'is_today': day_date == today,
                    'lessons': lessons,
                })

    today_kun = ['Dushanba','Seshanba','Chorshanba','Payshanba','Juma','Shanba','Yakshanba'][today.weekday()]
    return render(request, 'jadval/mobile.html', {
        'guruhlar': guruhlar,
        'selected_guruh': selected_guruh,
        'week_schedule': week_schedule,
        'weeks': weeks,
        'selected_week': selected_week,
        'today': today,
        'today_kun': today_kun,
    })


# ─── KIOSK ────────────────────────────────────────────────

def kiosk_home(request):
    guruhlar = Guruh.objects.all()
    selected_guruh = None
    week_schedule = []
    today_lessons = []
    weeks = []
    selected_week = None
    today = date.today()
    view_mode = request.GET.get('mode', 'haftalik')  # 'haftalik' | 'kunlik'

    guruh_id = request.GET.get('guruh')
    if guruh_id:
        selected_guruh = get_object_or_404(Guruh, id=guruh_id)
        weeks = selected_guruh.get_all_weeks()

        # Today's lessons
        today_lessons = list(Dars.objects.filter(guruh=selected_guruh, sana=today).order_by('para', 'vaqt'))

        # Weekly view
        week_param = request.GET.get('hafta')
        if week_param:
            try:
                selected_week = datetime.strptime(week_param, '%Y-%m-%d').date()
            except:
                selected_week = None
        if not selected_week:
            cur_mon = today - timedelta(days=today.weekday())
            selected_week = cur_mon if cur_mon in weeks else (weeks[0] if weeks else None)

        if selected_week:
            for i in range(6):
                day_date = selected_week + timedelta(days=i)
                lessons = list(Dars.objects.filter(guruh=selected_guruh, sana=day_date).order_by('para', 'vaqt'))
                week_schedule.append({
                    'date': day_date,
                    'day_name': HAFTA_KUNLARI[i],
                    'is_today': day_date == today,
                    'lessons': lessons,
                })

    today_kun = ['Dushanba','Seshanba','Chorshanba','Payshanba','Juma','Shanba','Yakshanba'][today.weekday()]
    return render(request, 'jadval/kiosk.html', {
        'guruhlar': guruhlar,
        'selected_guruh': selected_guruh,
        'week_schedule': week_schedule,
        'today_lessons': today_lessons,
        'weeks': weeks,
        'selected_week': selected_week,
        'today': today,
        'today_kun': today_kun,
        'view_mode': view_mode,
    })


def display(request):
    """Landscape sidebar kiosk — /display/ URL"""
    guruhlar = Guruh.objects.all()
    selected_guruh = None
    week_schedule = []
    today_lessons = []
    weeks = []
    selected_week = None
    today = date.today()
    view_mode = request.GET.get('mode', 'haftalik')

    guruh_id = request.GET.get('guruh')
    if guruh_id:
        selected_guruh = get_object_or_404(Guruh, id=guruh_id)
        weeks = selected_guruh.get_all_weeks()
        today_lessons = list(Dars.objects.filter(guruh=selected_guruh, sana=today).order_by('para', 'vaqt'))

        week_param = request.GET.get('hafta')
        if week_param:
            try:
                selected_week = datetime.strptime(week_param, '%Y-%m-%d').date()
            except:
                selected_week = None
        if not selected_week:
            cur_mon = today - timedelta(days=today.weekday())
            selected_week = cur_mon if cur_mon in weeks else (weeks[0] if weeks else None)

        if selected_week:
            for i in range(6):
                day_date = selected_week + timedelta(days=i)
                lessons = list(Dars.objects.filter(guruh=selected_guruh, sana=day_date).order_by('para', 'vaqt'))
                week_schedule.append({
                    'date': day_date,
                    'day_name': HAFTA_KUNLARI[i],
                    'is_today': day_date == today,
                    'lessons': lessons,
                })

    today_kun = ['Dushanba','Seshanba','Chorshanba','Payshanba','Juma','Shanba','Yakshanba'][today.weekday()]
    return render(request, 'jadval/display.html', {
        'guruhlar': guruhlar,
        'selected_guruh': selected_guruh,
        'week_schedule': week_schedule,
        'today_lessons': today_lessons,
        'weeks': weeks,
        'selected_week': selected_week,
        'today': today,
        'today_kun': today_kun,
        'view_mode': view_mode,
    })


# ─── ADMIN ────────────────────────────────────────────────

def admin_login(request):
    if request.user.is_authenticated:
        return redirect('admin_panel')
    if request.method == 'POST':
        user = authenticate(request, username=request.POST.get('username'), password=request.POST.get('password'))
        if user:
            login(request, user)
            return redirect('admin_panel')
        messages.error(request, "Login yoki parol noto'g'ri!")
    return render(request, 'jadval/admin_login.html')


def admin_logout(request):
    logout(request)
    return redirect('admin_login')


@login_required(login_url='/admin-panel/login/')
def admin_panel(request):
    guruhlar = Guruh.objects.all()
    total_dars = sum(g.darslar.count() for g in guruhlar)
    return render(request, 'jadval/admin_panel.html', {'guruhlar': guruhlar, 'total_dars': total_dars})


@login_required(login_url='/admin-panel/login/')
def guruh_create(request):
    ICONS = ['🎓','⚽','🏀','🏐','🎾','🥋','🤾','🚴','🧠','👔','🏫','📚','🏊','🏋️','🤸','⛹️','🎯','🏆','🎽','🏇','🥊','🏌️','🎱','🏓','🥅']
    COLORS = ['#00c6ff','#0072ff','#00e5a0','#f5c518','#ff6b6b','#a78bfa','#fb923c','#34d399','#f472b6','#60a5fa','#facc15','#e879f9']
    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        if not nom:
            messages.error(request, "Guruh nomi bo'sh bo'lmasligi kerak!")
            return render(request, 'jadval/guruh_form.html', {'icons': ICONS, 'colors': COLORS})
        guruh = Guruh.objects.create(
            nom=nom,
            rang=request.POST.get('rang', '#00c6ff'),
            ikonka=request.POST.get('ikonka', '🎓'),
            tartib=request.POST.get('tartib', 0)
        )
        if request.FILES.get('ikonka_rasm'):
            guruh.ikonka_rasm = request.FILES['ikonka_rasm']
            guruh.save()
        messages.success(request, f"'{nom}' guruhi yaratildi! Endi Excel jadval yuklang.")
        return redirect('guruh_excel_upload', pk=guruh.pk)
    return render(request, 'jadval/guruh_form.html', {'icons': ICONS, 'colors': COLORS})


@login_required(login_url='/admin-panel/login/')
def guruh_edit(request, pk):
    guruh = get_object_or_404(Guruh, pk=pk)
    ICONS = ['🎓','⚽','🏀','🏐','🎾','🥋','🤾','🚴','🧠','👔','🏫','📚','🏊','🏋️','🤸','⛹️','🎯','🏆','🎽','🏇','🥊','🏌️','🎱','🏓','🥅']
    COLORS = ['#00c6ff','#0072ff','#00e5a0','#f5c518','#ff6b6b','#a78bfa','#fb923c','#34d399','#f472b6','#60a5fa','#facc15','#e879f9']
    if request.method == 'POST':
        guruh.nom = request.POST.get('nom', '').strip()
        guruh.rang = request.POST.get('rang', '#00c6ff')
        guruh.ikonka = request.POST.get('ikonka', '🎓')
        guruh.tartib = request.POST.get('tartib', 0)
        if request.FILES.get('ikonka_rasm'):
            guruh.ikonka_rasm = request.FILES['ikonka_rasm']
        if request.POST.get('ikonka_rasm_ochir'):
            guruh.ikonka_rasm = None
        guruh.save()
        messages.success(request, "Guruh yangilandi!")
        return redirect('admin_panel')
    return render(request, 'jadval/guruh_form.html', {'guruh': guruh, 'icons': ICONS, 'colors': COLORS})


@login_required(login_url='/admin-panel/login/')
def guruh_delete(request, pk):
    guruh = get_object_or_404(Guruh, pk=pk)
    if request.method == 'POST':
        nom = guruh.nom
        guruh.delete()
        messages.success(request, f"'{nom}' o'chirildi.")
    return redirect('admin_panel')


@login_required(login_url='/admin-panel/login/')
def guruh_excel_upload(request, pk):
    guruh = get_object_or_404(Guruh, pk=pk)
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Fayl tanlanmadi!")
            return redirect('guruh_excel_upload', pk=pk)
        if not (excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls')):
            messages.error(request, "Faqat .xlsx yoki .xls fayl!")
            return redirect('guruh_excel_upload', pk=pk)

        suffix = '.xlsx' if excel_file.name.endswith('.xlsx') else '.xls'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            parsed = parse_excel(tmp_path)
            sheet_data = next((s for s in parsed if s.get('lessons')), None)
            if not sheet_data:
                messages.error(request, "Excel faylda dars ma'lumotlari topilmadi!")
                return redirect('guruh_excel_upload', pk=pk)

            Dars.objects.filter(guruh=guruh).delete()
            Dars.objects.bulk_create([Dars(guruh=guruh, **l) for l in sheet_data['lessons']])
            messages.success(request, f"✅ {len(sheet_data['lessons'])} ta dars yuklandi! ('{sheet_data['sheet_name']}' varag'idan)")
            if len(parsed) > 1:
                messages.info(request, f"💡 Faylda {len(parsed)} varaq bor edi — birinchi ma'lumotli varaq olindi.")
        except Exception as e:
            messages.error(request, f"Xato: {e}")
        finally:
            os.unlink(tmp_path)
        return redirect('admin_panel')

    return render(request, 'jadval/excel_upload.html', {
        'guruh': guruh,
        'lesson_count': Dars.objects.filter(guruh=guruh).count(),
    })


# ─── BULK UPLOAD ───────────────────────────────────────────

SHEET_TO_GROUP = {
    'рахбарлар':  {"nom": "Rahbarlar (Direktorlar)",         "ikonka": "👔", "rang": "#38bdf8"},
    'рах-ўрин':   {"nom": "Rahbar o'rinbosarlari",           "ikonka": "👔", "rang": "#0072ff"},
    'спорт псих': {"nom": "Sport psixologlari",              "ikonka": "🧠", "rang": "#a78bfa"},
    'кураш':      {"nom": "Kurash",                          "ikonka": "🥋", "rang": "#fbbf24"},
    'юнон рум':   {"nom": "Yunon-Rim kurashi",               "ikonka": "🥋", "rang": "#fb923c"},
    'гандбол':    {"nom": "Gandbol",                         "ikonka": "🤾", "rang": "#34d399"},
    'теннис-1':   {"nom": "Tennis-1",                        "ikonka": "🎾", "rang": "#34d399"},
    'теннис-2':   {"nom": "Tennis-2",                        "ikonka": "🎾", "rang": "#60a5fa"},
    'велоспорт':  {"nom": "Velosiped sporti",                "ikonka": "🚴", "rang": "#f472b6"},
    'ОТМ СФ':     {"nom": "OTM Sport faoliyati",             "ikonka": "🏆", "rang": "#facc15"},
    'мактаб-1':   {"nom": "Maktab-1",                        "ikonka": "🏫", "rang": "#f87171"},
    'мактаб-2':   {"nom": "Maktab-2",                        "ikonka": "🏫", "rang": "#fb923c"},
    'мактаб-3':   {"nom": "Maktab-3",                        "ikonka": "🏫", "rang": "#a78bfa"},
    'мактаб-4':   {"nom": "Maktab-4",                        "ikonka": "🏫", "rang": "#38bdf8"},
    'мактаб-5':   {"nom": "Maktab-5",                        "ikonka": "🏫", "rang": "#34d399"},
    'mtt+':       {"nom": "MTT+",                            "ikonka": "🎯", "rang": "#fbbf24"},
    'тарбиячи':   {"nom": "Tarbiyachilar",                   "ikonka": "📚", "rang": "#60a5fa"},
    'футбол-1':   {"nom": "Futbol-1",                        "ikonka": "⚽", "rang": "#34d399"},
    'футбол-2':   {"nom": "Futbol-2",                        "ikonka": "⚽", "rang": "#0072ff"},
    'футбол-3':   {"nom": "Futbol-3",                        "ikonka": "⚽", "rang": "#f87171"},
}

SKIP_SHEETS = ['хафталик', 'Лист1', 'Лист2', 'Лист4', 'МАРТ TAQSIMOT']


@login_required(login_url='/admin-panel/login/')
def bulk_upload(request):
    """Bitta Excel fayldan barcha guruhlarni yuklash."""
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Fayl tanlanmadi!")
            return redirect('bulk_upload')

        if not (excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls')):
            messages.error(request, "Faqat .xlsx yoki .xls fayl!")
            return redirect('bulk_upload')

        suffix = '.xlsx' if excel_file.name.endswith('.xlsx') else '.xls'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            from openpyxl import load_workbook
            import pandas as pd
            from .excel_parser import _parse_sheet

            wb = load_workbook(tmp_path, data_only=True)
            sheets = [s for s in wb.sheetnames if s not in SKIP_SHEETS]

            created_groups = 0
            total_lessons = 0
            errors = []

            for idx, sheet_name in enumerate(sheets):
                try:
                    df = pd.read_excel(tmp_path, sheet_name=sheet_name, header=None)
                    lessons = _parse_sheet(df)
                    if not lessons:
                        errors.append(f"'{sheet_name}': dars topilmadi")
                        continue

                    # Get group info from mapping or auto-create
                    info = SHEET_TO_GROUP.get(sheet_name, {
                        "nom": sheet_name,
                        "ikonka": "🎓",
                        "rang": "#38bdf8",
                    })

                    guruh = Guruh.objects.create(
                        nom=info["nom"],
                        ikonka=info["ikonka"],
                        rang=info["rang"],
                        tartib=idx + 1,
                    )
                    Dars.objects.bulk_create([Dars(guruh=guruh, **l) for l in lessons])
                    created_groups += 1
                    total_lessons += len(lessons)

                except Exception as e:
                    errors.append(f"'{sheet_name}': {str(e)[:60]}")

            messages.success(request,
                f"✅ {created_groups} ta guruh yaratildi, {total_lessons} ta dars yuklandi!")
            if errors:
                for err in errors:
                    messages.warning(request, f"⚠️ {err}")

        except Exception as e:
            messages.error(request, f"Fayl o'qishda xato: {e}")
        finally:
            os.unlink(tmp_path)

        return redirect('admin_panel')

    import json
    schema = [
        ("A — Sana",      "Dars sanasi (Excel sana formati, masalan 02.03.2026)"),
        ("B — Hafta kuni","Dushanba, Seshanba, ..."),
        ("C — Para",      "I, II, III (rim raqam)"),
        ("D — Vaqt",      "09.00-10.20 ko'rinishida"),
        ("E — Modul",     "Modul nomi (masalan: 2.3.Rahbarlik mahorati)"),
        ("F — Tur",       "nazariy / amaliy / ko'chma mashg'ulot"),
        ("G — O'qituvchi","Professor-o'qituvchi F.I.Sh., ilmiy darajasi"),
        ("H — Xona",      "Auditoriya (masalan: 15 aud)"),
    ]
    known_sheets = [{"sheet": k, "nom": v["nom"]} for k, v in SHEET_TO_GROUP.items()]
    return render(request, 'jadval/bulk_upload.html', {
        'schema': schema,
        'group_map': SHEET_TO_GROUP,
        'known_sheets_json': json.dumps(known_sheets, ensure_ascii=False),
    })


@login_required(login_url='/admin-panel/login/')
def clear_all(request):
    """Barcha guruh va darslarni o'chirish."""
    if request.method == 'POST':
        count = Guruh.objects.count()
        Guruh.objects.all().delete()  # cascades to Dars
        messages.success(request, f"✅ {count} ta guruh va barcha darslar o'chirildi.")
    return redirect('admin_panel')
